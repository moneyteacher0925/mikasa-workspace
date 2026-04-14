import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "견적서"

# 스타일 정의
title_font = Font(name='맑은 고딕', size=18, bold=True)
header_font = Font(name='맑은 고딕', size=10, bold=True)
normal_font = Font(name='맑은 고딕', size=10)
small_font = Font(name='맑은 고딕', size=9)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
right = Alignment(horizontal='right', vertical='center')

# 열 너비
ws.column_dimensions['A'].width = 6    # 순번
ws.column_dimensions['B'].width = 18   # 품목
ws.column_dimensions['C'].width = 10   # 규격
ws.column_dimensions['D'].width = 8    # 수량
ws.column_dimensions['E'].width = 10   # 단가
ws.column_dimensions['F'].width = 12   # 공급가액
ws.column_dimensions['G'].width = 10   # 세액
ws.column_dimensions['H'].width = 14   # 비고

# 제목
ws.merge_cells('A1:H1')
ws['A1'] = '견  적  서'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# 수신
ws.merge_cells('A3:H3')
ws['A3'] = '주식회사 갑조네 귀하'
ws['A3'].font = Font(name='맑은 고딕', size=11, bold=True)

# 견적번호/일자
ws.merge_cells('A4:D4')
ws['A4'] = 'NO. 20260403-D001'
ws['A4'].font = normal_font
ws.merge_cells('E4:H4')
ws['E4'] = '견적일자: 2026-04-03'
ws['E4'].font = normal_font
ws['E4'].alignment = Alignment(horizontal='right')

ws.merge_cells('A5:H5')
ws['A5'] = '견적요청에 감사드리며 아래와 같이 견적합니다.'
ws['A5'].font = normal_font

# 공급자 정보
ws.merge_cells('A7:H7')
ws['A7'] = '【공급자 정보】'
ws['A7'].font = header_font

info = [
    ('등록번호', '672-88-02456'),
    ('상호', '가든잇(주)'),
    ('성명', '권현우'),
    ('주소', '경기도 용인시 기흥구 동백중앙로 191 (중동, 씨티프라자) 8층 씨819호'),
    ('업태', '도매 및 소매업'),
    ('종목', '종합도매업, 전자상거래 소매업'),
    ('전화번호', '070-4589-9195'),
    ('팩스번호', '050-3837-92914'),
]

for i, (label, value) in enumerate(info):
    row = 8 + i
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'C{row}:H{row}')
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = normal_font

# 합계금액
row_total_label = 17
ws.merge_cells(f'A{row_total_label}:H{row_total_label}')
ws[f'A{row_total_label}'] = '합계금액: 일금 삼만일천구백 원정 (₩31,900) (부가세포함)'
ws[f'A{row_total_label}'].font = Font(name='맑은 고딕', size=11, bold=True)

# 테이블 헤더
header_row = 19
headers = ['순번', '품목', '규격', '수량', '단가', '공급가액', '세액', '비고']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border
    cell.fill = header_fill

# 데이터
items = [
    (1, '질석', '1.5L', 1, 800, 800, 80, '20개입'),
    (2, '펄라이트', '1.5L', 1, 800, 800, 80, '20개입'),
    (3, '훈탄', '1.5L', 1, 800, 800, 80, '20개입'),
    (4, '코코피트', '1.5L', 1, 800, 800, 80, '20개입'),
    (5, '피트모스', '1.5L', 1, 800, 800, 80, '20개입'),
    (6, '바크', '1.5L', 1, 700, 700, 70, '20개입'),
    (7, '적옥토', '1.5L', 1, 2100, 2100, 210, '10개입'),
    (8, '녹소토', '1.5L', 1, 1800, 1800, 180, '10개입'),
    (9, '휴가토', '1.5L', 1, 1000, 1000, 100, '10개입'),
    (10, '동생사', '1.5L', 1, 2100, 2100, 210, '10개입'),
    (11, '산야초', '1.5L', 1, 1800, 1800, 180, '10개입'),
    (12, '에스라이트', '1.5L', 1, 3200, 3200, 320, '10개입'),
    (13, '오키아타바크', '1.5L', 1, 2100, 2100, 210, '10개입'),
    (14, '다육이흙', '1.5L', 1, 1000, 1000, 100, '10개입'),
    (15, '화산석', '1.5L', 1, 1800, 1800, 180, '10개입'),
    (16, '마사토', '1.5L', 1, 800, 800, 80, '10개입'),
    (17, '세척마사토', '1.5L', 1, 1100, 1100, 110, '10개입'),
    (18, '분갈이흙', '1.5L', 1, 700, 700, 70, '20개입'),
    (19, '다육이흙', '8L', 1, 4800, 4800, 480, '1P=180개'),
]

for i, item in enumerate(items):
    row = header_row + 1 + i
    for col, val in enumerate(item, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if col in (1, 3, 4):
            cell.alignment = center
        elif col in (5, 6, 7):
            cell.alignment = right
            cell.number_format = '#,##0'
        elif col == 2:
            cell.alignment = left
        else:
            cell.alignment = center

# 합계 행
sum_row = header_row + 1 + len(items)
ws.cell(row=sum_row, column=1, value='합계').font = header_font
ws.cell(row=sum_row, column=1).alignment = center
ws.cell(row=sum_row, column=1).border = thin_border
for col in range(2, 5):
    ws.cell(row=sum_row, column=col).border = thin_border
ws.cell(row=sum_row, column=6, value=29000).font = header_font
ws.cell(row=sum_row, column=6).alignment = right
ws.cell(row=sum_row, column=6).border = thin_border
ws.cell(row=sum_row, column=6).number_format = '#,##0'
ws.cell(row=sum_row, column=7, value=2900).font = header_font
ws.cell(row=sum_row, column=7).alignment = right
ws.cell(row=sum_row, column=7).border = thin_border
ws.cell(row=sum_row, column=7).number_format = '#,##0'
ws.cell(row=sum_row, column=8).border = thin_border

# 비고 섹션
note_row = sum_row + 2
notes = [
    '【출고정보】',
    '- 발주단위: 파렛트',
    '- 출고방법: 파렛트상차 - 퀵화물',
    '',
    '【특이사항】',
    '- 재고상황/거리/배차상황 등에 따라 원하시는 일정에 수령이 어려울 수 있습니다.',
    '- 시즌: 재입고까지 최소 2주이상 소요 / 비시즌: 재입고까지 최소 1주이상 소요',
    '',
    '【운임비】',
    '- 거리상 예상운임비: 1.4톤 단독(2P) 80,000원 // 3.5톤 단독(4P) 115,000원 // 5톤 단독(10P) 140,000원',
    '- 지불방법: 상차지 선불',
]

for i, note in enumerate(notes):
    ws.merge_cells(f'A{note_row+i}:H{note_row+i}')
    cell = ws[f'A{note_row+i}']
    cell.value = note
    if note.startswith('【'):
        cell.font = header_font
    else:
        cell.font = small_font

# 하단 정보
bottom_row = note_row + len(notes) + 1
bottom_info = [
    ('납기일자', '발주후 2주 이내'),
    ('납품장소', '경기도 고양시 일산서구 대화로 362'),
    ('유효일자', '견적일로부터 30일'),
    ('결제조건', '협의'),
]

for i, (label, value) in enumerate(bottom_info):
    row = bottom_row + i
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'C{row}:H{row}')
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = normal_font

# 인쇄 설정
ws.print_area = f'A1:H{bottom_row + 3}'
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_A4

output_path = '/Users/kwoneren/.openclaw/workspace/gardeneat-gabzone-quote.xlsx'
wb.save(output_path)
print(f'저장 완료: {output_path}')
