import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "견적서"

# 스타일
title_font = Font(name='맑은 고딕', size=18, bold=True)
header_font = Font(name='맑은 고딕', size=10, bold=True)
normal_font = Font(name='맑은 고딕', size=10)
small_font = Font(name='맑은 고딕', size=9)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right', vertical='center')

# 열 너비
widths = {'A': 6, 'B': 16, 'C': 8, 'D': 10, 'E': 14, 'F': 14, 'G': 12, 'H': 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# === 제목 ===
ws.merge_cells('A1:H1')
ws['A1'] = '견  적  서'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# 수신
ws.merge_cells('A3:H3')
ws['A3'] = '주식회사 갑조네 귀하'
ws['A3'].font = Font(name='맑은 고딕', size=11, bold=True)

# 견적번호/일자
ws.merge_cells('A4:D4')
ws['A4'] = 'NO. 20260414-T001'
ws['A4'].font = normal_font
ws.merge_cells('E4:H4')
ws['E4'] = '견적일자: 2026-04-14'
ws['E4'].font = normal_font
ws['E4'].alignment = Alignment(horizontal='right')

ws.merge_cells('A5:H5')
ws['A5'] = '견적요청에 감사드리며 아래와 같이 견적합니다.'
ws['A5'].font = normal_font

# 공급자 정보
ws.merge_cells('A7:H7')
ws['A7'] = '【공급자 정보】'
ws['A7'].font = header_font

info = [
    ('등록번호', '672-88-02456'),
    ('상호', '가든잇(주)'),
    ('성명', '권현우'),
    ('주소', '경기도 용인시 기흥구 동백중앙로 191 (중동, 씨티프라자) 8층 씨819호'),
    ('업태', '도매 및 소매업'),
    ('종목', '종합도매업, 전자상거래 소매업'),
    ('전화번호', '070-4589-9195'),
    ('팩스번호', '050-3837-92914'),
]
for i, (label, value) in enumerate(info):
    row = 8 + i
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'C{row}:H{row}')
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = normal_font

# 품목: 2026 독일토분 변경 가격표
ws.merge_cells('A17:H17')
ws['A17'] = '2026 독일토분 변경 가격표 (부가세 별도)'
ws['A17'].font = Font(name='맑은 고딕', size=12, bold=True)
ws['A17'].alignment = center

# 테이블 헤더
header_row = 19
headers = ['순번', '품명', '크기', '단가(원)', '1박스당 수량', '1박스당 가격(원)', '비고', '']
col_headers = ['순번', '품명', '크기', '단가(원)', '1박스당 수량', '1박스당 가격(원)', '비고']
for col_idx, h in enumerate(col_headers, 1):
    cell = ws.cell(row=header_row, column=col_idx, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border
    cell.fill = header_fill

# 데이터
items = [
    # 일반형(토기)
    ('일반형(토기)', 4, 320, 1920, 614400),
    ('일반형(토기)', 5, 380, 900, 342000),
    ('일반형(토기)', 6, 435, 450, 195750),
    ('일반형(토기)', 7, 490, 300, 147000),
    ('일반형(토기)', 8, 590, 252, 148680),
    ('일반형(토기)', 9, 635, 160, 101600),
    ('일반형(토기)', 10, 725, 105, 76125),
    ('일반형(토기)', 11, 910, 84, 76440),
    ('일반형(토기)', 12, 1000, 72, 72000),
    ('일반형(토기)', 13, 1180, 45, 53100),
    ('일반형(토기)', 15, 1820, 28, 50960),
    ('일반형(토기)', 18, 2360, 18, 42480),
    ('일반형(토기)', 20, 3900, 12, 46800),
    # 일반형 받침
    ('일반형 받침', 7, 360, 840, 302400),
    ('일반형 받침', 8, 450, 560, 252000),
    ('일반형 받침', 9, 540, 340, 183600),
    ('일반형 받침', 10, 630, 304, 191520),
    ('일반형 받침', 11, 720, 192, 138240),
    ('일반형 받침', 13, 910, 126, 114660),
    ('일반형 받침', 15, 1180, 72, 84960),
    ('일반형 받침', 18, 1730, 55, 95150),
    ('일반형 받침', 22, 3550, 36, 127800),
    # 바솔트
    ('바솔트', 9, 810, 160, 129600),
    ('바솔트', 11, 1090, 72, 78480),
    ('바솔트', 13, 1450, 45, 65250),
    ('바솔트', 15, 1900, 28, 53200),
    ('바솔트', 18, 2720, 18, 48960),
    ('바솔트', 20, 5000, 12, 60000),
    # 그래니티
    ('그래니티', 11, 1540, 72, 110880),
    ('그래니티', 13, 2090, 45, 94050),
    ('그래니티', 15, 3180, 28, 89040),
    ('그래니티', 18, 4540, 18, 81720),
    ('그래니티', 20, 6400, 12, 76800),
    # 컬러받침
    ('컬러받침', 9, 990, 460, 455400),
    ('컬러받침', 11, 1180, 240, 283200),
    ('컬러받침', 13, 1360, 144, 195840),
    ('컬러받침', 14, 1900, 112, 212800),
    ('컬러받침', 16, 2360, 84, 198240),
    ('컬러받침', 18, 3000, 60, 180000),
    ('컬러받침', 20, 4000, 33, 132000),
    ('컬러받침', 22, 5000, 36, 180000),
]

data_row = header_row + 1
current_category = ''
seq = 0
for item in items:
    category, size, price, qty, box_price = item
    seq += 1
    row = data_row
    
    ws.cell(row=row, column=1, value=seq).font = normal_font
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=category).font = normal_font
    ws.cell(row=row, column=2).alignment = left_al
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=size).font = normal_font
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4, value=price).font = normal_font
    ws.cell(row=row, column=4).alignment = right_al
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).number_format = '#,##0'
    
    ws.cell(row=row, column=5, value=qty).font = normal_font
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).number_format = '#,##0'
    
    ws.cell(row=row, column=6, value=box_price).font = normal_font
    ws.cell(row=row, column=6).alignment = right_al
    ws.cell(row=row, column=6).border = thin_border
    ws.cell(row=row, column=6).number_format = '#,##0'
    
    ws.cell(row=row, column=7).font = normal_font
    ws.cell(row=row, column=7).border = thin_border
    
    # 카테고리별 색상
    if category in ('바솔트', '그래니티', '컬러받침'):
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = light_fill
    
    data_row += 1

# 비고 섹션
note_row = data_row + 1
notes = [
    '【거래조건】',
    '- 상기 가격은 부가세 별도 금액입니다.',
    '- 발주단위: 1박스 단위',
    '- 운임: 가든잇 화물출고 (가든잇 부담)',
    '- 결제조건: 납품일 기준 익월 말일까지',
    '',
    '【특이사항】',
    '- 재고상황에 따라 원하시는 일정에 수령이 어려울 수 있습니다.',
    '- 시즌: 재입고까지 최소 2주 이상 소요 / 비시즌: 최소 1주 이상 소요',
]

for i, note in enumerate(notes):
    ws.merge_cells(f'A{note_row+i}:G{note_row+i}')
    cell = ws[f'A{note_row+i}']
    cell.value = note
    if note.startswith('【'):
        cell.font = header_font
    else:
        cell.font = small_font

# 하단 정보
bottom_row = note_row + len(notes) + 1
bottom_info = [
    ('납기일자', '발주후 2주 이내'),
    ('유효일자', '견적일로부터 30일'),
    ('결제조건', '납품일 기준 익월 말일까지'),
]
for i, (label, value) in enumerate(bottom_info):
    row = bottom_row + i
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'C{row}:G{row}')
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = normal_font

# 인쇄 설정
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_A4

output_path = '/Users/kwoneren/.openclaw/workspace/gardeneat-gabzone-tobunn-quote.xlsx'
wb.save(output_path)
print(f'저장 완료: {output_path}')
