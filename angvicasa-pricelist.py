import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "앙비까사 토분 리스트"

title_font = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
header_font = Font(name='맑은 고딕', size=10, bold=True)
normal_font = Font(name='맑은 고딕', size=10)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
header_font_w = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
alt_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
center = Alignment(horizontal='center', vertical='center')
right_al = Alignment(horizontal='right', vertical='center')
left_al = Alignment(horizontal='left', vertical='center')

widths = {'A': 6, 'B': 16, 'C': 12, 'D': 24, 'E': 14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# 제목
ws.merge_cells('A1:E1')
ws['A1'] = '2026년 앙비까사 토분 리스트'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
ws.row_dimensions[1].height = 35

# 헤더
headers = ['순번', '이름', '수량', '색상', '가격(원)']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font_w
    cell.alignment = center
    cell.border = thin_border
    cell.fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')

# 데이터
items = [
    # BSC-A 시리즈
    ('BSC-A 60', '440', 'R(레드)', 270),
    ('BSC-A 110', '70', 'P(피치)', 1100),
    ('BSC-A 130', '48', 'P(피치)', 1800),
    ('BSC-A 150', '36', 'P(피치),BG(베이지)', 2200),
    ('BSC-A 170', '24', 'P(피치),BG(베이지)', 2500),
    ('BSC-A 190', '18', 'P(피치),BG(베이지)', 2700),
    ('BSC-A 210', '12', 'P(피치),BG(베이지)', 3000),
    ('BSC-A 230', '9', 'P(피치),BG(베이지)', 5000),
    # BSC-T 시리즈
    ('BSC-T 101', '90', 'R(레드)', 1000),
    ('BSC-T 101', '90', 'BG(베이지)', 1300),
    ('BSC-T 150', '24', 'P(피치)', 2700),
    ('BSC-T 170', '18', 'P(피치)', 4200),
    ('BSC-T 200', '12', 'P(피치)', 5200),
    ('BSC-T 250', '6', 'P(피치)', 10500),
    ('BSC-T 300', '4', 'P(피치)', 16000),
    ('BSC-T 342', '2', 'P(피치)', 27000),
    # CON.L 시리즈
    ('CON.L 140', '32', 'P(피치),BG(베이지)', 3200),
    ('CON.L 170', '18', 'P(피치),BG(베이지)', 4800),
    ('CON.L 170', '18 SET', 'C(초코)', 7500),
    ('CON.L 200', '12', 'P(피치),BG(베이지)', 5800),
    ('CON.L 230', '6', 'P(피치),BG(베이지)', 7500),
    ('CON.L 230', '5 SET', 'C(초코)', 11500),
    ('CON.L 260', '6', 'P(피치),BG(베이지)', 9500),
    ('CON.L 290', '4', 'P(피치),BG(베이지)', 16000),
    ('CON.L 340', '4', 'P(피치),BG(베이지)', 21000),
    # APC 시리즈
    ('APC 160', '24', 'P(피치),BG(베이지)', 3200),
    ('APC 200', '12', 'P(피치),BG(베이지)', 6500),
    ('APC 250', '6', 'P(피치),BG(베이지)', 9500),
    ('APC 310', '4', 'P(피치)', 12000),
    # OLV
    ('OLV 160', '18', 'P(피치)', 6500),
    # BWL.B 시리즈
    ('BWL.B 250', '16', 'P(피치),BG(베이지)', 5800),
    ('BWL.B 300', '8', 'P(피치),BG(베이지)', 8500),
    ('BWL.B 350', '6', 'P(피치)', 12000),
    # BWL.C 시리즈
    ('BWL.C 200', '24', 'P(피치),BG(베이지)', 3200),
    ('BWL.C 250', '16', 'P(피치),BG(베이지)', 5800),
    # BWL.G
    ('BWL.G 200', '18 SET', 'C(초코)', 11500),
    # DTR-D 시리즈
    ('DTR-D50', '500,450', 'R(레드)', 270),
    ('DTR-D50', '500,450', 'BG(베이지)', 380),
    ('DTR-D60', '360', 'R(레드)', 300),
    ('DTR-D60', '360', 'BG(베이지)', 420),
    ('DTR-D70', '240', 'R(레드)', 350),
    ('DTR-D70', '240', 'BG(베이지)', 500),
    ('DTR-D80', '180', 'BG(베이지)', 550),
    ('DTR-D80', '180', 'R(레드)', 450),
    ('DTR-D90', '120', 'R(레드)', 550),
    ('DTR-D90', '120', 'BG(베이지)', 850),
    ('DTR-D100', '90', 'R(레드)', 650),
    ('DTR-D100', '90', 'BG(베이지)', 950),
    # DTR-D 대형
    ('DTR-D110', '75', 'P(피치),BG(베이지)', 1100),
    ('DTR-D130', '48', 'P(피치),BG(베이지)', 2000),
    ('DTR-D150', '36', 'P(피치),BG(베이지)', 2500),
    ('DTR-D170', '24', 'P(피치),BG(베이지)', 3000),
    ('DTR-D190', '18,24', 'P(피치),W(화이트)', 4300),
    ('DTR-D210', '12', 'P(피치),W(화이트)', 5300),
    ('DTR-D230', '9', 'P(피치),W(화이트)', 6300),
    # BLL O 시리즈
    ('BLL O 140', '30', 'P(피치),BG(베이지)', 3200),
    ('BLL O 170', '18', 'P(피치),BG(베이지)', 4200),
    ('BLL O 170', '21 SET', 'C(초코)', 6500),
    ('BLL O 210', '9', 'P(피치),BG(베이지)', 6500),
    ('BLL O 210', '8 SET', 'C(초코)', 10000),
    ('BLL O 250', '6', 'P(피치),BG(베이지)', 10500),
    # HAR 시리즈
    ('HAR.182', '6', 'P(피치),BG(베이지)', 10500),
    ('HAR.182', '4 SET', 'C(초코)', 13000),
    # DR 시리즈
    ('DR-320', '4', 'P(피치),BG(베이지)', 16000),
    ('DR-360', '3,2', 'P(피치),BG(베이지)', 25000),
    # RND.R 시리즈
    ('RND.R 130', '45 SET', 'C(초코)', 4000),
    ('RND.R 160', '21 SET', 'C(초코)', 6000),
    ('RND.R 190', '12 SET', 'C(초코)', 9000),
    # ITA 시리즈
    ('ITA 220', '8', 'P(피치),W(화이트)', 5500),
    ('ITA 160', '24', 'P(피치),W(화이트)', 2500),
    ('ITA 130', '45', 'P(피치),W(화이트)', 1800),
    # HM.HORN 시리즈
    ('HM.HORN 170', '24 SET', 'P(피치)', 5000),
    ('HM.HORN 200', '12 SET', 'P(피치)', 7000),
    ('HM.HORN 240', '12 SET', 'P(피치)', 11000),
]

for i, (name, qty, color, price) in enumerate(items):
    row = 4 + i
    ws.cell(row=row, column=1, value=i+1).font = normal_font
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=name).font = normal_font
    ws.cell(row=row, column=2).alignment = left_al
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=qty).font = normal_font
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4, value=color).font = normal_font
    ws.cell(row=row, column=4).alignment = left_al
    ws.cell(row=row, column=4).border = thin_border
    
    ws.cell(row=row, column=5, value=price).font = normal_font
    ws.cell(row=row, column=5).alignment = right_al
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).number_format = '#,##0'
    
    if i % 2 == 1:
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = alt_fill

ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_A4

output = '/Users/kwoneren/.openclaw/workspace/angvicasa-pricelist-2026.xlsx'
wb.save(output)
print(f'저장 완료: {output}')
