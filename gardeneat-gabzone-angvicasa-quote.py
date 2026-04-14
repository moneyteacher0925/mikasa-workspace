import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "견적서"

title_font = Font(name='맑은 고딕', size=18, bold=True)
header_font = Font(name='맑은 고딕', size=10, bold=True)
normal_font = Font(name='맑은 고딕', size=10)
small_font = Font(name='맑은 고딕', size=9)
red_font = Font(name='맑은 고딕', size=10, color='FF0000')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right', vertical='center')

widths = {'A': 6, 'B': 10, 'C': 8, 'D': 8, 'E': 12, 'F': 35, 'G': 10, 'H': 12, 'I': 14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# 제목
ws.merge_cells('A1:I1')
ws['A1'] = '견  적  서'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# 수신
ws.merge_cells('A3:I3')
ws['A3'] = '주식회사 갑조네 귀하'
ws['A3'].font = Font(name='맑은 고딕', size=11, bold=True)

ws.merge_cells('A4:D4')
ws['A4'] = 'NO. 20260414-V001'
ws['A4'].font = normal_font
ws.merge_cells('E4:I4')
ws['E4'] = '견적일자: 2026-04-14'
ws['E4'].font = normal_font
ws['E4'].alignment = Alignment(horizontal='right')

ws.merge_cells('A5:I5')
ws['A5'] = '견적요청에 감사드리며 아래와 같이 견적합니다.'
ws['A5'].font = normal_font

# 공급자 정보
ws.merge_cells('A7:I7')
ws['A7'] = '【공급자 정보】'
ws['A7'].font = header_font

info = [
    ('등록번호', '672-88-02456'),
    ('상호', '가든잇(주)'),
    ('성명', '권현우'),
    ('주소', '경기도 용인시 기흥구 동백중앙로 191 (중동, 씨티프라자) 8층 씨819호'),
    ('업태', '도매 및 소매업'),
    ('종목', '종합도매업, 전자상거래 소매업'),
    ('전화번호', '070-4589-9195'),
]
for i, (label, value) in enumerate(info):
    row = 8 + i
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'C{row}:I{row}')
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = normal_font

# 제목
ws.merge_cells('A16:I16')
ws['A16'] = '2026 앙비까사 변경 가격표 (부가세 별도)'
ws['A16'].font = Font(name='맑은 고딕', size=12, bold=True)
ws['A16'].alignment = center

# 테이블 헤더
hr = 18
col_headers = ['순번', '품명', '크기', '색상', '코드', '품명(상세)', '단가(원)', '1박스당 수량', '1박스당 가격(원)']
for col_idx, h in enumerate(col_headers, 1):
    cell = ws.cell(row=hr, column=col_idx, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border
    cell.fill = header_fill

items = [
    ('DTR', 50, 'R', 'G01177', '도토리화분(5cm)피치(레드)', 240, 500, 120000),
    ('DTR', 50, 'G', 'G01176', '도토리화분(5cm)베이지', 340, 500, 170000),
    ('DTR', 60, 'R', 'G99868', '도토리화분(6cm) 토기', 270, 360, 97200),
    ('DTR', 60, 'G', 'G99867', '도토리화분(6cm) 베이지', 380, 360, 136800),
    ('DTR', 70, 'R', 'G99883', '도토리화분(7cm) 토기', 320, 240, 76800),
    ('DTR', 70, 'G', 'G99882', '도토리화분(7cm) 베이지', 460, 240, 110400),
    ('DTR', 80, 'R', 'G99885', '도토리화분(8cm) 토기', 410, 180, 73800),
    ('DTR', 80, 'G', 'G99884', '도토리화분(8cm) 베이지', 500, 180, 90000),
    ('DTR', 90, 'R', 'G99887', '도토리화분(9cm) 토기', 500, 120, 60000),
    ('DTR', 90, 'G', 'G99886', '도토리화분(9cm) 베이지', 780, 120, 93600),
    ('DTR', 110, 'P/G', 'G00140', '도토리화분(소) (11cm)', 1000, 75, 75000),
    ('DTR', 130, 'P/G', 'G00141', '도토리화분(중) (13cm)', 1810, 48, 86880),
    ('DTR', 150, 'P/G', 'G00142', '도토리화분(대) (15cm)', 2270, 36, 81720),
    ('BLL', 140, 'P/G', 'G01705', '앙비까사 진주 토분(베이지)/(소)(14cm)', 2900, 30, 87000),
    ('BLL', 170, 'P/G', 'G01666', '앙비까사 진주 토분(피치)/(중)(17cm)', 3800, 18, 68400),
    ('BLL', 210, 'P/G', 'G01665', '앙비까사 진주 토분(베이지)/(대)(21cm)', 5900, 9, 53100),
    ('CON', 140, 'P/G', 'G00146', '앙비까사 콘형 토분(피치)(소)', 2900, 32, 92800),
    ('CON', 170, 'P/G', 'G00147', '앙비까사 콘형 토분(피치)(중)', 4350, 18, 78300),
    ('MDN', 110, 'P/G', 'G01649', '앙비까사 빈 토분(베이지) / 사이즈:11호', '', '', ''),
    ('MDN', 140, 'P/G', 'G01650', '앙비까사 빈 토분(베이지) / 사이즈:14호', '', '', ''),
    ('MDN', 160, 'P/G', 'G01651', '앙비까사 빈 토분(베이지) / 사이즈:16호', '', '', ''),
    ('CLD.R', 120, 'P/G', 'G100340', '앙비까사 구름토분 12호 CLD.R120 P', '', '', ''),
    ('CLD.R', 150, 'P/G', 'G100341', '앙비까사 구름토분 15호 CLD.R150 P', '', '', ''),
    ('CLD.R', 180, 'P/G', 'G100342', '앙비까사 구름토분 17호 CLD.R170 P', '', '', ''),
    ('HM.HEEL', 120, 'P/G', 'G01662', '앙비까사 힐스 토분(피치)/소+받침', '', '', ''),
    ('HM.HEEL', 140, 'P/G', 'G01663', '앙비까사 힐스 토분(피치)/중+받침', '', '', ''),
]

data_row = hr + 1
for i, item in enumerate(items):
    row = data_row + i
    name, size, color, code, desc, price, qty, box_price = item
    
    ws.cell(row=row, column=1, value=i+1).font = normal_font
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=name).font = normal_font
    ws.cell(row=row, column=2).alignment = center
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=size).font = normal_font
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4, value=color).font = normal_font
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=4).border = thin_border
    
    ws.cell(row=row, column=5, value=code).font = normal_font
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).border = thin_border
    
    ws.cell(row=row, column=6, value=desc).font = normal_font
    ws.cell(row=row, column=6).alignment = left_al
    ws.cell(row=row, column=6).border = thin_border
    
    if price != '':
        ws.cell(row=row, column=7, value=price).font = normal_font
        ws.cell(row=row, column=7).number_format = '#,##0'
    ws.cell(row=row, column=7).alignment = right_al
    ws.cell(row=row, column=7).border = thin_border
    
    if qty != '':
        ws.cell(row=row, column=8, value=qty).font = normal_font
        ws.cell(row=row, column=8).number_format = '#,##0'
    ws.cell(row=row, column=8).alignment = center
    ws.cell(row=row, column=8).border = thin_border
    
    if box_price != '':
        ws.cell(row=row, column=9, value=box_price).font = red_font
        ws.cell(row=row, column=9).number_format = '#,##0'
    ws.cell(row=row, column=9).alignment = right_al
    ws.cell(row=row, column=9).border = thin_border

# 비고
note_row = data_row + len(items) + 1
notes = [
    '【거래조건】',
    '- 상기 가격은 부가세 별도 금액입니다.',
    '- 발주단위: 1박스 단위',
    '- 운임: 가든잇 화물출고 (가든잇 부담)',
    '- 결제조건: 납품일 기준 익월 말일까지',
    '',
    '【특이사항】',
    '- 재고상황에 따라 원하시는 일정에 수령이 어려울 수 있습니다.',
    '- 시즌: 재입고까지 최소 2주 이상 소요 / 비시즌: 최소 1주 이상 소요',
    '- 빈칸 항목은 단가 확인 후 업데이트 예정',
]
for i, note in enumerate(notes):
    ws.merge_cells(f'A{note_row+i}:I{note_row+i}')
    cell = ws[f'A{note_row+i}']
    cell.value = note
    cell.font = header_font if note.startswith('【') else small_font

bottom_row = note_row + len(notes) + 1
for i, (label, value) in enumerate([
    ('납기일자', '발주후 2주 이내'),
    ('유효일자', '견적일로부터 30일'),
    ('결제조건', '납품일 기준 익월 말일까지'),
]):
    row = bottom_row + i
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'C{row}:I{row}')
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = normal_font

ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4

output = '/Users/kwoneren/.openclaw/workspace/gardeneat-gabzone-angvicasa-quote.xlsx'
wb.save(output)
print(f'저장 완료: {output}')
